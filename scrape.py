from bs4 import BeautifulSoup
import requests
import openpyxl
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title='Top Rated Movies'
print(excel.sheetnames)
sheet.append(['movie rank','movie name','release year','rating'])

source = requests.get('https://www.imdb.com/chart/top/')
soup = BeautifulSoup(source.text, 'html.parser')
movies = soup.find('tbody',class_ ='lister-list').find_all('tr')
for movie in movies:
    name = movie.find('td',class_ ='titleColumn').a.text
    rank = movie.find('td',class_='titleColumn').get_text(strip=True).split( '.')[0]
    year = movie.find('td',class_='titleColumn').span.text.strip('()')
    rating = movie.find('td',class_='ratingColumn imdbRating').strong.text
    print(name,rank,year,rating)
    sheet.append([rank,name,year,rating])
    excel.save('Top movies of imdb.xlsx')